#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para actualizar el archivo Excel con datos de PurchaseOrderItem.

Compara productos del PurchaseOrder con la columna "ref:" (SKU) del Excel
y actualiza la columna L con quantity_units.
"""
import os
import sys
import django

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.local')
django.setup()

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("❌ Error: openpyxl no está instalado")
    print("Instala con: pip install --break-system-packages openpyxl")
    sys.exit(1)

from purchase_orders.models import PurchaseOrder, PurchaseOrderItem


def update_excel_from_purchase_order(purchase_order_id, excel_file='miquel.xlsx', output_file=None):
    """
    Actualiza el Excel con datos del PurchaseOrder usando SKU.
    
    Args:
        purchase_order_id: ID del PurchaseOrder
        excel_file: Ruta al archivo Excel
        output_file: Archivo de salida (si None, sobreescribe el original)
    """
    
    print(f"\n{'='*70}")
    print(f"ACTUALIZANDO EXCEL CON DATOS DE PURCHASE ORDER #{purchase_order_id}")
    print(f"{'='*70}\n")
    
    # 1. Obtener el PurchaseOrder y sus items
    try:
        po = PurchaseOrder.objects.prefetch_related('items__product').select_related('market', 'provider').get(pk=purchase_order_id)
    except PurchaseOrder.DoesNotExist:
        print(f"❌ Error: No existe PurchaseOrder con ID {purchase_order_id}")
        return False
    
    print(f"📦 Purchase Order: {po}")
    print(f"   Proveedor: {po.provider.name}")
    print(f"   Mercado: {po.market.name if po.market else 'Sin mercado'}")
    print(f"   Status: {po.status}")
    print(f"   Items: {po.items.count()}\n")
    
    # Crear diccionario de productos {sku: quantity_units}
    po_items = {}
    for item in po.items.all():
        product_sku = item.product.sku
        po_items[product_sku] = {
            'quantity_units': item.quantity_units,
            'product_id': item.product.id,
            'product_name': item.product.name,
        }
    
    print(f"📋 Productos en el Purchase Order:")
    for sku, data in po_items.items():
        print(f"   • SKU {sku}: {data['product_name']} - {data['quantity_units']} unidades")
    
    # 2. Cargar el archivo Excel
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"\n❌ Error: No se encontró el archivo {excel_file}")
        return False
    except Exception as e:
        print(f"\n❌ Error al abrir Excel: {e}")
        return False
    
    print(f"\n📄 Archivo Excel: {excel_file}")
    print(f"   Hoja activa: {ws.title}")
    print(f"   Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    # 3. Encontrar la columna "ref:" EXACTAMENTE (buscar en fila 2)
    ref_col = None
    header_row = 2  # El encabezado está en la fila 2
    data_start_row = 4  # Los datos empiezan en la fila 4
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value and str(cell_value).strip().lower() == "ref:":
            ref_col = col
            break
    
    if not ref_col:
        print("\n❌ Error: No se encontró la columna 'ref:' en la fila 2")
        print("\nColumnas encontradas en fila 2:")
        for col in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(header_row, col).value
            if val:
                print(f"   Columna {chr(64+col)}: '{val}'")
        return False
    
    print(f"\n🔍 Columna 'ref:' (SKU) encontrada: Columna {ref_col} ({chr(64+ref_col)})")
    print(f"   Columna L (destino): Columna 12")
    print(f"   Fila de encabezados: {header_row}")
    print(f"   Fila inicio de datos: {data_start_row}")
    
    # 4. LIMPIAR COLUMNAS F-M (columnas 6 a 13) - Datos de mercados
    print(f"\n🧼 Limpiando columnas F-M (mercados)...")
    cleaned = 0
    for row in range(data_start_row, ws.max_row + 1):
        for col in range(6, 14):  # Columnas F(6) hasta M(13)
            ws.cell(row, col).value = None
            cleaned += 1
    print(f"   ✓ {cleaned} celdas limpiadas (columnas F-M)")
    
    # 5. Comparar y actualizar
    print(f"\n{'='*70}")
    print("COMPARANDO Y ACTUALIZANDO...")
    print(f"{'='*70}\n")
    
    updated = 0
    not_found = []
    
    for row in range(data_start_row, ws.max_row + 1):  # Empezar desde fila 4 (datos)
        ref_value = ws.cell(row, ref_col).value
        
        if not ref_value:
            continue
        
        # Convertir a string y limpiar (puede venir como número o string)
        sku = str(ref_value).strip()
        
        # Buscar coincidencia exacta por SKU
        if sku in po_items:
            quantity = po_items[sku]['quantity_units']
            product_name = po_items[sku]['product_name']
            
            # Actualizar columna L (columna 12)
            cell = ws.cell(row, 12)
            cell.value = quantity
            
            # Aplicar formato: centrado y fuente estándar
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"✓ Fila {row}: SKU {sku} ({product_name}) → {quantity} unidades")
            updated += 1
        else:
            # SKU no encontrado en el Purchase Order
            not_found.append({'row': row, 'sku': sku})
            print(f"✗ Fila {row}: SKU {sku} (no está en el Purchase Order)")
    
    # 6. Guardar el archivo
    output_path = output_file or excel_file
    try:
        wb.save(output_path)
        print(f"\n{'='*70}")
        print(f"✅ Archivo guardado exitosamente: {output_path}")
        print(f"{'='*70}")
    except Exception as e:
        print(f"\n❌ Error al guardar archivo: {e}")
        return False
    
    # 7. Resumen
    print(f"\n📊 RESUMEN:")
    print(f"   ✓ Actualizados: {updated} productos")
    print(f"   ✗ No encontrados: {len(not_found)} productos")
    
    if not_found:
        print(f"\n❌ PRODUCTOS NO ENCONTRADOS EN EL PURCHASE ORDER:")
        for item in not_found:
            print(f"   Fila {item['row']}: SKU '{item['sku']}'")
    
    print(f"\n{'='*70}\n")
    return True


if __name__ == "__main__":
    # Verificar que se proporcionó el ID del PurchaseOrder como argumento
    if len(sys.argv) < 2:
        print("\n❌ Error: Debes proporcionar el ID del PurchaseOrder como argumento\n")
        print("Uso: python update_excel_from_po.py <purchase_order_id>\n")
        print("Ejemplo: python update_excel_from_po.py 123\n")
        
        # Mostrar órdenes disponibles como ayuda
        print("📦 PURCHASE ORDERS DISPONIBLES:\n")
        orders = PurchaseOrder.objects.select_related('provider').all()[:20]
        for order in orders:
            items_count = order.items.count()
            print(f"   ID {order.id}: {order.provider.name} - {order.status} ({items_count} items)")
        print()
        sys.exit(1)
    
    # Obtener el ID del PurchaseOrder
    try:
        po_id = int(sys.argv[1])
    except ValueError:
        print("❌ Error: El ID del PurchaseOrder debe ser un número")
        sys.exit(1)
    
    # Ejecutar actualización
    success = update_excel_from_purchase_order(
        purchase_order_id=po_id,
        excel_file='miquel.xlsx',
        output_file='miquel_actualizado.xlsx'  # Guardar en archivo nuevo
    )
    
    if success:
        print("✅ Proceso completado exitosamente!\n")
    else:
        print("❌ El proceso terminó con errores\n")
        sys.exit(1)
