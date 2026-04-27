"""Utilities to export purchase orders to Excel and PDF."""

from io import BytesIO
from collections import defaultdict
from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image, HRFlowable, PageBreak,
)

COLOR_DARK = "0F172A"
COLOR_GREEN = "10B981"
COLOR_ORANGE = "F97316"
LOGO_PATH = "/app/media/organization_logos/kacha_logo.png"
MARKETS_PER_PAGE = 10


def _safe_text(value, default=""):
    return str(value if value is not None else default)


def _order_market_name(order):
    return _safe_text(getattr(getattr(order, "market", None), "name", None), "Sin tienda")


def _provider_name(order):
    return _safe_text(getattr(getattr(order, "provider", None), "name", None), "Proveedor")


def _get_org(order):
    try:
        return order.market.organization if order.market else None
    except Exception:
        return None


def _get_org_data(org):
    return {
        "name": _safe_text(getattr(org, "name", ""), "Kacha Digital BCN"),
        "address": _safe_text(getattr(org, "address", ""), ""),
        "cif": _safe_text(getattr(org, "cif", ""), ""),
        "phone": _safe_text(getattr(org, "contact_phone", ""), ""),
        "email": _safe_text(getattr(org, "contact_email", ""), ""),
    }


def _get_barcode(product):
    if not product:
        return ""
    try:
        bc = product.barcodes.filter(is_primary=True).first() or product.barcodes.first()
        return _safe_text(getattr(bc, "code", None), "") if bc else ""
    except Exception:
        return ""


def _grouped_rows_from_orders(orders):
    rows_by_order = []
    consolidated = defaultdict(lambda: {
        "product_name": "", "sku": "", "barcode": "",
        "markets": defaultdict(int), "total_units": 0
    })
    market_names = []
    seen_markets = set()

    for order in orders:
        market_name = _order_market_name(order)
        if market_name not in seen_markets:
            seen_markets.add(market_name)
            market_names.append(market_name)

        order_items = list(order.items.select_related("product").prefetch_related("product__barcodes").all().order_by("product__name"))
        detail_items = []

        for item in order_items:
            product = getattr(item, "product", None)
            product_name = _safe_text(getattr(product, "name", None), "Producto")
            sku = _safe_text(getattr(product, "sku", None), "")
            barcode = _get_barcode(product)
            qty = int(item.quantity_units or 0)

            detail_items.append({
                "product_name": product_name,
                "sku": sku,
                "barcode": barcode,
                "quantity_units": qty,
                "purchase_unit": _safe_text(getattr(item, "purchase_unit", "boxes"), "boxes"),
            })

            key = item.product_id
            consolidated[key]["product_name"] = product_name
            consolidated[key]["sku"] = sku
            consolidated[key]["barcode"] = barcode
            consolidated[key]["markets"][market_name] += qty
            consolidated[key]["total_units"] += qty

        rows_by_order.append({
            "order_id": order.id,
            "market_name": market_name,
            "provider_name": _provider_name(order),
            "items": detail_items,
        })

    consolidated_rows = sorted([
        {
            "product_name": v["product_name"],
            "sku": v["sku"],
            "barcode": v["barcode"],
            "markets": dict(v["markets"]),
            "total_units": v["total_units"],
        }
        for v in consolidated.values()
    ], key=lambda x: x["product_name"].lower())

    market_names.sort(key=lambda x: x.lower())
    return rows_by_order, consolidated_rows, market_names


def build_grouped_purchase_order_excel(orders):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    org = _get_org(orders[0]) if orders else None
    o = _get_org_data(org)
    provider_name = _provider_name(orders[0]) if orders else "Proveedor"
    _, consolidated_rows, market_names = _grouped_rows_from_orders(orders)

    dark_fill = PatternFill(fill_type="solid", fgColor=COLOR_DARK)
    orange_fill = PatternFill(fill_type="solid", fgColor=COLOR_ORANGE)
    green_fill = PatternFill(fill_type="solid", fgColor=COLOR_GREEN)
    light_fill = PatternFill(fill_type="solid", fgColor="F1F5F9")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    amber_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")

    # Logo
    ws.row_dimensions[1].height = 60
    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    max_col = len(market_names) + 4  # producto + sku + barcode + markets + total
    col_letter = get_column_letter(max_col)

    ws.merge_cells(f"B1:{col_letter}1")
    ws["B1"].value = o["name"].upper()
    ws["B1"].font = Font(bold=True, color=COLOR_DARK, size=16)
    ws["B1"].alignment = Alignment(vertical="center")

    ws.merge_cells(f"B2:{col_letter}2")
    ws["B2"].value = o["address"]
    ws["B2"].font = Font(color="64748B", size=10)

    ws.merge_cells(f"B3:{col_letter}3")
    ws["B3"].value = f"CIF: {o['cif']}  |  Tel: {o['phone']}  |  {o['email']}"
    ws["B3"].font = Font(color="64748B", size=10)

    ws.merge_cells(f"A5:{col_letter}5")
    ws["A5"].fill = orange_fill
    ws.row_dimensions[5].height = 4

    ws.merge_cells(f"A7:{col_letter}7")
    ws["A7"].value = f"PEDIDO CONSOLIDADO — {provider_name.upper()}"
    ws["A7"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A7"].fill = dark_fill
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 30

    ws.merge_cells(f"A8:{col_letter}8")
    ws["A8"].value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Pedidos: {len(orders)}  |  Tiendas: {len(market_names)}"
    ws["A8"].font = Font(color="64748B", size=10)
    ws["A8"].alignment = Alignment(horizontal="center")

    # Headers
    header_row = 10
    headers = ["Producto", "SKU", "Cód. Barras"] + market_names + ["TOTAL"]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = orange_fill if value == "TOTAL" else dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 28

    # Ancho columnas
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    for i in range(4, len(market_names) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(len(market_names) + 4)].width = 12

    # Datos
    data_row = header_row + 1
    for i, item in enumerate(consolidated_rows):
        fill = light_fill if i % 2 == 0 else white_fill
        ws.cell(row=data_row, column=1, value=item["product_name"]).fill = fill
        ws.cell(row=data_row, column=2, value=item["sku"]).fill = fill
        ws.cell(row=data_row, column=3, value=item["barcode"]).fill = fill
        for idx, market_name in enumerate(market_names, start=4):
            qty = int(item["markets"].get(market_name, 0))
            cell = ws.cell(row=data_row, column=idx, value=qty)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            if qty > 0:
                cell.font = Font(bold=True, color=COLOR_GREEN, size=11)
        total_cell = ws.cell(row=data_row, column=len(market_names) + 4, value=int(item["total_units"]))
        total_cell.fill = amber_fill
        total_cell.font = Font(bold=True, color="92400E", size=11)
        total_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[data_row].height = 22
        data_row += 1

    # Totals
    totals_row = data_row
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(row=totals_row, column=1).fill = dark_fill
    ws.cell(row=totals_row, column=2).fill = dark_fill
    ws.cell(row=totals_row, column=3).fill = dark_fill
    for idx, market_name in enumerate(market_names, start=4):
        col_total = sum(int(r["markets"].get(market_name, 0)) for r in consolidated_rows)
        cell = ws.cell(row=totals_row, column=idx, value=col_total)
        cell.fill = dark_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center")
    grand_total = sum(int(r["total_units"]) for r in consolidated_rows)
    gt_cell = ws.cell(row=totals_row, column=len(market_names) + 4, value=grand_total)
    gt_cell.fill = orange_fill
    gt_cell.font = Font(bold=True, color="FFFFFF", size=13)
    gt_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[totals_row].height = 28

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_grouped_purchase_order_pdf(orders):
    out = BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=landscape(A4),
        leftMargin=1.2*cm, rightMargin=1.2*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm
    )
    styles = getSampleStyleSheet()
    elements = []

    org = _get_org(orders[0]) if orders else None
    o = _get_org_data(org)
    provider_name = _provider_name(orders[0]) if orders else "Proveedor"
    _, consolidated_rows, market_names = _grouped_rows_from_orders(orders)

    def build_header():
        header_data = [[]]
        if os.path.exists(LOGO_PATH):
            try:
                header_data[0].append(Image(LOGO_PATH, width=2*cm, height=2*cm))
            except Exception:
                header_data[0].append("")
        else:
            header_data[0].append("")

        company_info = (
            f'<b><font size="13" color="#0F172A">{o["name"].upper()}</font></b><br/>'
            f'<font size="8" color="#64748B">{o["address"]}</font><br/>'
            f'<font size="8" color="#64748B">CIF: {o["cif"]}  |  Tel: {o["phone"]}  |  {o["email"]}</font>'
        )
        header_data[0].append(Paragraph(company_info, styles["Normal"]))

        doc_info = (
            f'<b><font size="11" color="#0F172A">PEDIDO CONSOLIDADO</font></b><br/>'
            f'<font size="8" color="#64748B">Proveedor: {provider_name}</font><br/>'
            f'<font size="8" color="#64748B">Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Pedidos: {len(orders)}</font>'
        )
        header_data[0].append(Paragraph(doc_info, styles["Normal"]))

        ht = Table(header_data, colWidths=[2.5*cm, 16*cm, 8*cm])
        ht.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return ht

    # Dividir tiendas en chunks de MARKETS_PER_PAGE
    market_chunks = [market_names[i:i+MARKETS_PER_PAGE] for i in range(0, max(len(market_names), 1), MARKETS_PER_PAGE)]

    for chunk_idx, chunk_markets in enumerate(market_chunks):
        if chunk_idx > 0:
            elements.append(PageBreak())

        elements.append(build_header())
        elements.append(Spacer(1, 0.25*cm))
        elements.append(HRFlowable(width="100%", thickness=3, color=colors.HexColor("#F97316")))
        elements.append(Spacer(1, 0.3*cm))

        # Calcular anchos de columna
        page_w = landscape(A4)[0] - 2.4*cm
        n = len(chunk_markets)
        fixed_cols = 3  # producto + sku + barcode
        total_col_w = 2.5*cm
        fixed_w = 2.5*cm + 2*cm + total_col_w  # sku + barcode + total
        product_col_w = min(7*cm, max(4*cm, page_w - fixed_w - (n * 2.5*cm)))
        market_col_w = (page_w - product_col_w - fixed_w) / max(n, 1)
        market_col_w = max(1.5*cm, market_col_w)
        col_widths = [product_col_w, 2*cm, 2.5*cm] + [market_col_w] * n + [total_col_w]

        # Tabla
        table_data = [["Producto", "SKU", "Cód. Barras"] + chunk_markets + ["TOTAL"]]
        for item in consolidated_rows:
            row = [item["product_name"], item.get("sku", ""), item.get("barcode", "")]
            for m in chunk_markets:
                row.append(int(item["markets"].get(m, 0)))
            row.append(int(item["total_units"]))
            table_data.append(row)

        totals_row = ["TOTAL", "", ""]
        for m in chunk_markets:
            totals_row.append(sum(int(r["markets"].get(m, 0)) for r in consolidated_rows))
        totals_row.append(sum(int(r["total_units"]) for r in consolidated_rows))
        table_data.append(totals_row)

        n_rows = len(table_data)

        row_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (-1, 0), (-1, 0), colors.HexColor("#F97316")),
            ("BACKGROUND", (0, n_rows-1), (-1, n_rows-1), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, n_rows-1), (-1, n_rows-1), colors.white),
            ("FONTNAME", (0, n_rows-1), (-1, n_rows-1), "Helvetica-Bold"),
            ("BACKGROUND", (-1, n_rows-1), (-1, n_rows-1), colors.HexColor("#F97316")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("ROWHEIGHT", (0, 0), (-1, -1), 20),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (0, -1), 6),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TEXTCOLOR", (3, 1), (-2, n_rows-2), colors.HexColor("#10B981")),
            ("FONTNAME", (3, 1), (-2, n_rows-2), "Helvetica-Bold"),
            ("TEXTCOLOR", (-1, 1), (-1, n_rows-2), colors.HexColor("#92400E")),
            ("BACKGROUND", (-1, 1), (-1, n_rows-2), colors.HexColor("#FEF3C7")),
        ]

        for i in range(1, n_rows - 1):
            if i % 2 == 0:
                row_styles.append(("BACKGROUND", (0, i), (-2, i), colors.HexColor("#F1F5F9")))

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(row_styles))
        elements.append(table)

    doc.build(elements)
    out.seek(0)
    return out


def build_purchase_order_excel(order):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    org = _get_org(order)
    o = _get_org_data(org)

    dark_fill = PatternFill(fill_type="solid", fgColor=COLOR_DARK)
    orange_fill = PatternFill(fill_type="solid", fgColor=COLOR_ORANGE)
    light_fill = PatternFill(fill_type="solid", fgColor="F1F5F9")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    ws.row_dimensions[1].height = 60
    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width = 80
            img.height = 80
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws.merge_cells("B1:F1")
    ws["B1"].value = o["name"].upper()
    ws["B1"].font = Font(bold=True, color=COLOR_DARK, size=16)
    ws["B1"].alignment = Alignment(vertical="center")

    ws.merge_cells("B2:F2")
    ws["B2"].value = o["address"]
    ws["B2"].font = Font(color="64748B", size=10)

    ws.merge_cells("B3:F3")
    ws["B3"].value = f"CIF: {o['cif']}  |  Tel: {o['phone']}  |  {o['email']}"
    ws["B3"].font = Font(color="64748B", size=10)

    ws.merge_cells("A5:F5")
    ws["A5"].fill = orange_fill
    ws.row_dimensions[5].height = 4

    ws["A7"] = "PEDIDO DE COMPRA"
    ws["A7"].font = Font(bold=True, color=COLOR_DARK, size=13)
    ws.merge_cells("A7:F7")

    info = [
        ("Pedido #", order.id),
        ("Proveedor", _provider_name(order)),
        ("Tienda", _order_market_name(order)),
        ("Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Estado", _safe_text(getattr(order, "status", ""), "")),
    ]
    row = 9
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="64748B", size=10)
        ws.cell(row=row, column=2, value=str(value)).font = Font(color=COLOR_DARK, size=10)
        row += 1

    header_row = row + 1
    headers = ["Producto", "SKU", "Cód. Barras", "Cantidad", "Unidad"]
    widths = [35, 14, 18, 12, 12]
    for col_idx, (value, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 28

    data_row = header_row + 1
    for i, item in enumerate(order.items.select_related("product").prefetch_related("product__barcodes").all().order_by("product__name")):
        product = getattr(item, "product", None)
        fill = light_fill if i % 2 == 0 else white_fill
        ws.cell(row=data_row, column=1, value=_safe_text(getattr(product, "name", None), "Producto")).fill = fill
        ws.cell(row=data_row, column=2, value=_safe_text(getattr(product, "sku", None), "")).fill = fill
        ws.cell(row=data_row, column=3, value=_get_barcode(product)).fill = fill
        qty_cell = ws.cell(row=data_row, column=4, value=int(item.quantity_units or 0))
        qty_cell.fill = fill
        qty_cell.alignment = Alignment(horizontal="center")
        qty_cell.font = Font(bold=True, color=COLOR_GREEN, size=11)
        ws.cell(row=data_row, column=5, value=_safe_text(getattr(item, "purchase_unit", "boxes"), "boxes")).fill = fill
        ws.row_dimensions[data_row].height = 22
        data_row += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_purchase_order_pdf(order):
    out = BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=landscape(A4),
        leftMargin=1.2*cm, rightMargin=1.2*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm
    )
    styles = getSampleStyleSheet()
    elements = []

    org = _get_org(order)
    o = _get_org_data(org)

    header_data = [[]]
    if os.path.exists(LOGO_PATH):
        try:
            header_data[0].append(Image(LOGO_PATH, width=2*cm, height=2*cm))
        except Exception:
            header_data[0].append("")
    else:
        header_data[0].append("")

    company_info = (
        f'<b><font size="13" color="#0F172A">{o["name"].upper()}</font></b><br/>'
        f'<font size="8" color="#64748B">{o["address"]}</font><br/>'
        f'<font size="8" color="#64748B">CIF: {o["cif"]}  |  Tel: {o["phone"]}  |  {o["email"]}</font>'
    )
    header_data[0].append(Paragraph(company_info, styles["Normal"]))

    order_info = (
        f'<b><font size="10" color="#0F172A">PEDIDO #{order.id}</font></b><br/>'
        f'<font size="8" color="#64748B">Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}</font><br/>'
        f'<font size="8" color="#64748B">Proveedor: {_provider_name(order)}</font><br/>'
        f'<font size="8" color="#64748B">Tienda: {_order_market_name(order)}</font>'
    )
    header_data[0].append(Paragraph(order_info, styles["Normal"]))

    ht = Table(header_data, colWidths=[2.5*cm, 16*cm, 8*cm])
    ht.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 0.25*cm))
    elements.append(HRFlowable(width="100%", thickness=3, color=colors.HexColor("#F97316")))
    elements.append(Spacer(1, 0.3*cm))

    data = [["Producto", "SKU", "Cód. Barras", "Cantidad", "Unidad"]]
    for item in order.items.select_related("product").prefetch_related("product__barcodes").all().order_by("product__name"):
        product = getattr(item, "product", None)
        data.append([
            _safe_text(getattr(product, "name", None), "Producto"),
            _safe_text(getattr(product, "sku", None), ""),
            _get_barcode(product),
            int(item.quantity_units or 0),
            _safe_text(getattr(item, "purchase_unit", "boxes"), "boxes"),
        ])

    table = Table(data, colWidths=[10*cm, 3*cm, 4*cm, 3*cm, 3*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("FONTNAME", (3, 1), (3, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#10B981")),
        ("FONTSIZE", (3, 1), (3, -1), 12),
        ("ROWHEIGHT", (0, 0), (-1, -1), 22),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (0, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F1F5F9"), colors.white]),
    ]))
    elements.append(table)

    doc.build(elements)
    out.seek(0)
    return out
