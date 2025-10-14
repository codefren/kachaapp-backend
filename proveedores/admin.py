from django.contrib import admin
from django.contrib import messages
from django import forms
import re
from simple_history.admin import SimpleHistoryAdmin
from django.urls import reverse
from django.utils.html import format_html

from .models import Provider, Product, ProductBarcode
from purchase_orders.models import PurchaseOrder


class ProviderUploadForm(forms.ModelForm):
    """Form para Provider con campo adicional para subir Excel de productos.

    Formato esperado (.xlsx): columnas con encabezados (no sensibles a mayúsculas):
    - sku (clave única)
    - name
    - units_per_box (opcional, entero)
    - amount_boxes (opcional, entero)
    - barcodes (opcional, lista separada por comas)
    """

    products_excel = forms.FileField(
        required=False,
        help_text=(
            "Sube un Excel (.xlsx) con columnas: sku, name, units_per_box, amount_boxes, barcodes (opc.). "
            "Se crearán/actualizarán productos y se asociarán a este proveedor."
        ),
        label="Carga de productos (Excel)",
    )

    class Meta:
        model = Provider
        fields = ("name", "order_deadline_time", "order_available_weekdays")


class PurchaseOrderInline(admin.TabularInline):
    model = PurchaseOrder
    extra = 0
    fields = ("status", "market", "ordered_by", "created_at")
    readonly_fields = ("created_at",)
    autocomplete_fields = ("market", "ordered_by")
    ordering = ("-created_at",)


@admin.register(Provider)
class ProviderAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "order_deadline_time", "get_weekdays_display", "last_po_id", "last_po_status", "pivot_link")
    search_fields = ("name",)
    ordering = ("name",)
    fields = ("name", "order_deadline_time", "order_available_weekdays", "products_excel", "pivot_link")
    form = ProviderUploadForm
    inlines = [PurchaseOrderInline]
    readonly_fields = ("pivot_link",)

    def get_weekdays_display(self, obj):
        """Muestra los días de la semana de forma legible."""
        if not obj.order_available_weekdays:
            return "Sin días configurados"

        weekday_names = {
            0: 'Lun', 1: 'Mar', 2: 'Mié', 3: 'Jue',
            4: 'Vie', 5: 'Sáb', 6: 'Dom'
        }

        days = [weekday_names.get(day, str(day)) for day in obj.order_available_weekdays]
        return ", ".join(days)

    get_weekdays_display.short_description = "Días disponibles"

    def last_po_id(self, obj):
        po = obj.purchase_orders.order_by("-created_at").first()
        return po.id if po else None
    last_po_id.short_description = "Última PO"

    def last_po_status(self, obj):
        po = obj.purchase_orders.order_by("-created_at").first()
        return po.status if po else "—"
    last_po_status.short_description = "Estado última PO"

    def pivot_link(self, obj):
        try:
            url = reverse("admin:purchaseorder_pivot") + f"?provider={obj.id}"
            return format_html('<a class="button" href="{}">Ver pivot de pedidos</a>', url)
        except Exception:
            return ""
    pivot_link.short_description = "Pivot pedidos (Admin)"

    def save_model(self, request, obj, form, change):
        """Procesa la carga de Excel si se proporcionó y guarda el Provider normalmente."""
        super().save_model(request, obj, form, change)

        upload = form.cleaned_data.get("products_excel") if hasattr(form, "cleaned_data") else None
        if not upload:
            return

        # Utilidades de normalización y mapeo flexible por regex
        def normalize_colname(c: str) -> str:
            return re.sub(r"\s+", " ", str(c).lower().strip())

        # Patrones más tolerantes (acentos, abreviaturas, puntos y espacios)
        COL_MAP_PATTERNS = {
            # "Cód. Barras", "Código de barras", "Cod Barras", etc.
            r"c[oó]d\.?\s*(de\s*)?barras?\b|barcode|ean|gtin": "barcodes",
            # sku o referencia: "ref", "ref.", "referencia", "codigo", "código"
            r"sku\b|ref\s*[:\.]?\b|referencia\b|c[oó]digo\b": "sku",
            # nombre / producto / descripción
            r"nombre\b|producto\b|descripci[óo]n\b|art[íi]culo\b": "name",
            # cajas / boxes
            r"x\s*cajas\b|cajas\b|boxes\b": "amount_boxes",
            # unidades por caja
            r"unid.*caja|unidades.*caja|units.*box|u/?caja|u\.*/caja": "units_per_box",
            # proveedor
            r"proveedor\b|supplier\b|vendor\b": "provider_name",
        }

        def detect_column_mapping(columns):
            mapping = {}
            for col in columns:
                low = normalize_colname(col)
                for pat, target in COL_MAP_PATTERNS.items():
                    if re.search(pat, low):
                        mapping[target] = col
                        break
            return mapping

        def find_header_row(ws, max_scan=30):
            """Busca la fila de encabezados en las primeras `max_scan` filas.
            Retorna (row_index, columns_list) o (None, None) si no encuentra.
            """
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
                if not row:
                    continue
                columns = [c if (c is not None and str(c).strip() != "") else f"col_{j+1}" for j, c in enumerate(row)]
                mapping = detect_column_mapping(columns)
                # Requerimos al menos "name" y uno entre "sku" o "barcodes"
                if mapping.get("name") and (mapping.get("sku") or mapping.get("barcodes")):
                    return i, columns
            return None, None

        # Procesar Excel
        try:
            try:
                import openpyxl  # type: ignore
            except Exception as e:  # pragma: no cover
                messages.error(
                    request,
                    "openpyxl no está instalado. Agrega 'openpyxl' a requirements e inténtalo de nuevo.",
                )
                return

            wb = openpyxl.load_workbook(upload, data_only=True)

            total_created, total_updated = 0, 0
            total_bc_created, total_bc_existing = 0, 0
            total_rows_scanned, total_rows_imported = 0, 0
            total_rows_skipped_no_identifier = 0

            # Iterar todas las hojas para soportar formatos por proveedor/hoja
            for ws in wb.worksheets:
                # Detectar fila de encabezados de forma flexible
                header_row_idx, columns = find_header_row(ws)
                if not header_row_idx:
                    # No se pudo detectar encabezado usable
                    continue
                mapping = detect_column_mapping(columns)

                # Validar al menos un identificador: sku o barcodes
                if not (mapping.get("sku") or mapping.get("barcodes")):
                    # No se puede identificar producto en esta hoja
                    continue

                # Índices por nombre de columna original
                name_to_index = {columns[i]: i for i in range(len(columns))}

                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    total_rows_scanned += 1
                    def val_for(target_key):
                        src_col = mapping.get(target_key)
                        if not src_col:
                            return None
                        idx = name_to_index.get(src_col)
                        if idx is None or idx >= len(row):
                            return None
                        return row[idx]

                    raw_name = val_for("name")
                    name = (str(raw_name).strip() if raw_name is not None else None) or "Unknown"

                    raw_sku = val_for("sku")
                    sku = str(raw_sku).strip() if raw_sku not in (None, "") else None

                    # Si no hay sku, tomar primer barcode como sku de fallback
                    raw_barcodes = val_for("barcodes")
                    barcode_list = []
                    if raw_barcodes not in (None, ""):
                        # Separadores comunes: coma o punto y coma
                        barcode_list = [b.strip() for b in re.split(r"[;,]", str(raw_barcodes)) if str(b).strip()]
                    if not sku and barcode_list:
                        sku = barcode_list[0]

                    # Si aún no tenemos identificador, saltar
                    if not sku:
                        total_rows_skipped_no_identifier += 1
                        continue

                    # Parseo de enteros tolerante
                    def parse_int(x, default=None):
                        try:
                            if x is None:
                                return default
                            s = str(x).strip()
                            if s == "":
                                return default
                            return int(float(s))  # maneja números venidos como float
                        except Exception:
                            return default

                    units_per_box = parse_int(val_for("units_per_box"))
                    amount_boxes = parse_int(val_for("amount_boxes"))

                    obj.refresh_from_db(fields=["id"])  # asegurar PK
                    product, was_created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            "name": name,
                            **({"units_per_box": units_per_box} if units_per_box is not None else {}),
                            **({"amount_boxes": amount_boxes} if amount_boxes is not None else {}),
                        },
                    )

                    # Asociar proveedor
                    product.providers.add(obj)
                    product.save()
                    total_created += 1 if was_created else 0
                    total_updated += 0 if was_created else 1
                    total_rows_imported += 1

                    # Barcodes opcionales
                    if barcode_list:
                        for i, code in enumerate(barcode_list):
                            if not code:
                                continue
                            pb, pb_created = ProductBarcode.objects.get_or_create(
                                code=code,
                                defaults={
                                    "product": product,
                                    "is_primary": (i == 0),
                                    "type": ProductBarcode.BarcodeType.EAN13,
                                },
                            )
                            if not pb_created:
                                # Asegurar relación con el producto si ya existía
                                if pb.product_id != product.id:
                                    pb.product = product
                                    pb.save(update_fields=["product"])
                            total_bc_created += 1 if pb_created else 0
                            total_bc_existing += 0 if pb_created else 1

            if total_created + total_updated == 0:
                messages.warning(
                    request,
                    (
                        "Excel procesado pero no se detectaron filas válidas de productos. "
                        f"Filas escaneadas: {total_rows_scanned}. Revise los encabezados (p. ej. 'Descripción', 'ref:', 'Cód. Barras') y que haya datos."
                    ),
                )
            else:
                messages.success(
                    request,
                    (
                        f"Excel procesado: {total_created} productos creados, {total_updated} actualizados. "
                        f"Códigos de barras: {total_bc_created} creados, {total_bc_existing} existentes. "
                        f"Filas importadas: {total_rows_imported}, omitidas sin identificador: {total_rows_skipped_no_identifier}."
                    ),
                )
        except Exception as e:  # pragma: no cover
            messages.error(request, f"Error procesando Excel: {e}")


class ProductBarcodeInline(admin.TabularInline):
    model = ProductBarcode
    extra = 1
    fields = ("code", "type", "is_primary", "notes")
    show_change_link = True
    classes = ("collapse",)


class HasBarcodeFilter(admin.SimpleListFilter):
    title = "Has barcode"
    parameter_name = "has_barcode"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        value = self.value()
        if value == "yes":
            return queryset.filter(barcodes__isnull=False).distinct()
        if value == "no":
            return queryset.filter(barcodes__isnull=True)
        return queryset


class HasPrimaryBarcodeFilter(admin.SimpleListFilter):
    title = "Has primary barcode"
    parameter_name = "has_primary_barcode"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        value = self.value()
        if value == "yes":
            return queryset.filter(barcodes__is_primary=True).distinct()
        if value == "no":
            return queryset.exclude(barcodes__is_primary=True).distinct()
        return queryset


@admin.register(Product)
class ProductAdmin(SimpleHistoryAdmin):
    list_display = ("id", "name", "sku", "units_per_box", "amount_boxes")
    search_fields = ("name", "sku", "providers__name", "barcodes__code")
    list_filter = ("providers", HasBarcodeFilter, HasPrimaryBarcodeFilter)
    ordering = ("name",)
    filter_horizontal = ("providers",)
    inlines = [ProductBarcodeInline]

    readonly_fields = ("image_preview",)
    fields = (
        "name",
        "sku",
        "providers",
        "units_per_box",
        "amount_boxes",
        "image",
        "image_preview",
    )

    def image_preview(self, obj):
        if obj and obj.image:
            return f'<img src="{obj.image.url}" style="max-height: 100px;" />'
        return ""
    image_preview.short_description = "Preview"
    image_preview.allow_tags = True


@admin.register(ProductBarcode)
class ProductBarcodeAdmin(admin.ModelAdmin):
    list_display = ("id", "code", "type", "is_primary", "product")
    search_fields = ("code", "product__name", "product__sku")
    list_filter = ("type", "is_primary")
    autocomplete_fields = ("product",)
    ordering = ("code",)

