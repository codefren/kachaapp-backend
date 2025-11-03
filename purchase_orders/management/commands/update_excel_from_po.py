# -*- coding: utf-8 -*-
"""
Comando Django para actualizar el archivo Excel con datos de PurchaseOrderItem.

Compara productos del PurchaseOrder con la columna "ref:" (SKU) del Excel
y actualiza la columna L con quantity_units.

Uso:
    python manage.py update_excel_from_po <purchase_order_id> [--excel FILENAME] [--output FILENAME]
"""
from django.core.management.base import BaseCommand, CommandError

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    raise ImportError(
        "openpyxl no está instalado. "
        "Instala con: pip install --break-system-packages openpyxl"
    )

from purchase_orders.models import PurchaseOrder


class Command(BaseCommand):
    help = 'Actualiza el archivo Excel con datos del PurchaseOrder usando SKU'

    def add_arguments(self, parser):
        # Argumento posicional requerido: ID del PurchaseOrder
        parser.add_argument(
            'purchase_order_id',
            type=int,
            help='ID del PurchaseOrder a procesar'
        )
        
        # Argumentos opcionales
        parser.add_argument(
            '--excel',
            type=str,
            default='miquel.xlsx',
            help='Ruta al archivo Excel de entrada (default: miquel.xlsx)'
        )
        
        parser.add_argument(
            '--output',
            type=str,
            default=None,
            help='Ruta al archivo Excel de salida (default: sobrescribe el original)'
        )

    def handle(self, *args, **options):
        purchase_order_id = options['purchase_order_id']
        excel_file = options['excel']
        output_file = options['output']
        
        self.stdout.write(f"\n{'='*70}")
        self.stdout.write(f"ACTUALIZANDO EXCEL CON DATOS DE PURCHASE ORDER #{purchase_order_id}")
        self.stdout.write(f"{'='*70}\n")
        
        # 1. Obtener el PurchaseOrder y sus items
        try:
            po = PurchaseOrder.objects.prefetch_related('items__product').select_related('market', 'provider').get(pk=purchase_order_id)
        except PurchaseOrder.DoesNotExist:
            raise CommandError(f'No existe PurchaseOrder con ID {purchase_order_id}')
        
        self.stdout.write(f"📦 Purchase Order: {po}")
        self.stdout.write(f"   Proveedor: {po.provider.name}")
        self.stdout.write(f"   Mercado: {po.market.name if po.market else 'Sin mercado'}")
        self.stdout.write(f"   Status: {po.status}")
        self.stdout.write(f"   Items: {po.items.count()}\n")
        
        # Crear diccionario de productos {sku: quantity_units}
        po_items = {}
        for item in po.items.all():
            product_sku = item.product.sku
            po_items[product_sku] = {
                'quantity_units': item.quantity_units,
                'product_id': item.product.id,
                'product_name': item.product.name,
            }
        
        self.stdout.write("📋 Productos en el Purchase Order:")
        for sku, data in po_items.items():
            self.stdout.write(f"   • SKU {sku}: {data['product_name']} - {data['quantity_units']} unidades")
        
        # 2. Cargar el archivo Excel
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except FileNotFoundError:
            raise CommandError(f"No se encontró el archivo {excel_file}")
        except Exception as e:
            raise CommandError(f"Error al abrir Excel: {e}")
        
        self.stdout.write(f"\n📄 Archivo Excel: {excel_file}")
        self.stdout.write(f"   Hoja activa: {ws.title}")
        self.stdout.write(f"   Filas: {ws.max_row}, Columnas: {ws.max_column}")
        
        # 3. Encontrar la columna "ref:" EXACTAMENTE (buscar en fila 2)
        ref_col = None
        header_row = 2  # El encabezado está en la fila 2
        data_start_row = 4  # Los datos empiezan en la fila 4
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col).value
            if cell_value and str(cell_value).strip().lower() == "ref:":
                ref_col = col
                break
        
        if not ref_col:
            self.stdout.write("\nColumnas encontradas en fila 2:")
            for col in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(header_row, col).value
                if val:
                    self.stdout.write(f"   Columna {chr(64+col)}: '{val}'")
            raise CommandError("No se encontró la columna 'ref:' en la fila 2")
        
        self.stdout.write(f"\n🔍 Columna 'ref:' (SKU) encontrada: Columna {ref_col} ({chr(64+ref_col)})")
        self.stdout.write(f"   Columna L (destino): Columna 12")
        self.stdout.write(f"   Fila de encabezados: {header_row}")
        self.stdout.write(f"   Fila inicio de datos: {data_start_row}")
        
        # 4. LIMPIAR COLUMNAS F-M (columnas 6 a 13) - Datos de mercados
        self.stdout.write(f"\n🧼 Limpiando columnas F-M (mercados)...")
        cleaned = 0
        for row in range(data_start_row, ws.max_row + 1):
            for col in range(6, 14):  # Columnas F(6) hasta M(13)
                ws.cell(row, col).value = None
                cleaned += 1
        self.stdout.write(f"   ✓ {cleaned} celdas limpiadas (columnas F-M)")
        
        # 5. Comparar y actualizar
        self.stdout.write(f"\n{'='*70}")
        self.stdout.write("COMPARANDO Y ACTUALIZANDO...")
        self.stdout.write(f"{'='*70}\n")
        
        updated = 0
        not_found = []
        
        for row in range(data_start_row, ws.max_row + 1):  # Empezar desde fila 4 (datos)
            ref_value = ws.cell(row, ref_col).value
            
            if not ref_value:
                continue
            
            # Convertir a string y limpiar (puede venir como número o string)
            sku = str(ref_value).strip()
            
            # Buscar coincidencia exacta por SKU
            if sku in po_items:
                quantity = po_items[sku]['quantity_units']
                product_name = po_items[sku]['product_name']
                
                # Actualizar columna L (columna 12)
                cell = ws.cell(row, 12)
                cell.value = quantity
                
                # Aplicar formato: centrado y fuente estándar
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                self.stdout.write(
                    self.style.SUCCESS(f"✓ Fila {row}: SKU {sku} ({product_name}) → {quantity} unidades")
                )
                updated += 1
            else:
                # SKU no encontrado en el Purchase Order
                not_found.append({'row': row, 'sku': sku})
                self.stdout.write(
                    self.style.WARNING(f"✗ Fila {row}: SKU {sku} (no está en el Purchase Order)")
                )
        
        # 6. Guardar el archivo
        output_path = output_file or excel_file
        try:
            wb.save(output_path)
            self.stdout.write(f"\n{'='*70}")
            self.stdout.write(self.style.SUCCESS(f"✅ Archivo guardado exitosamente: {output_path}"))
            self.stdout.write(f"{'='*70}")
        except Exception as e:
            raise CommandError(f"Error al guardar archivo: {e}")
        
        # 7. Resumen
        self.stdout.write(f"\n📊 RESUMEN:")
        self.stdout.write(f"   ✓ Actualizados: {updated} productos")
        self.stdout.write(f"   ✗ No encontrados: {len(not_found)} productos")
        
        if not_found:
            self.stdout.write(f"\n❌ PRODUCTOS NO ENCONTRADOS EN EL PURCHASE ORDER:")
            for item in not_found:
                self.stdout.write(f"   Fila {item['row']}: SKU '{item['sku']}'")
        
        self.stdout.write(f"\n{'='*70}\n")
        self.stdout.write(self.style.SUCCESS('✅ Proceso completado exitosamente!\n'))
